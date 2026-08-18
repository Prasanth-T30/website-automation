"""Excel and PDF exports of the payment ledger.

Both take the same already-filtered rows, so a download always matches what
the Finance screen was showing when the button was pressed — exporting the
whole ledger when the user is looking at one college would be a quiet lie.

Amounts are written to Excel as real numbers, not preformatted strings: the
point of an .xlsx over a .csv is that the recipient can sum a column.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.payment import PaymentTransaction
from app.models.student import Student

BRAND_BLUE = "3569AC"
BRAND_BLUE_RGB = (53, 105, 172)
BRAND_TEAL_RGB = (21, 181, 184)
BRAND_DARK_RGB = (20, 20, 20)
BRAND_GREY_RGB = (74, 74, 74)

COMPANY_NAME = "DVein Innovations Pvt. Ltd."

_HEADERS = ["Receipt", "Student", "College", "Amount", "Method", "Date", "Recorded by"]

# fpdf2's built-in fonts are latin-1 only. Typographic punctuation and any
# name outside that range would raise mid-render, so PDF text goes through
# here first. Excel is UTF-8 and needs none of this.
_PDF_SUBSTITUTIONS = {
    "—": "-",  # em dash
    "–": "-",  # en dash
    "‘": "'", "’": "'",
    "“": '"', "”": '"',
    "₹": "Rs. ",  # rupee sign
    "·": "-",  # middle dot
}


def _pdf_safe(text: str) -> str:
    for bad, good in _PDF_SUBSTITUTIONS.items():
        text = text.replace(bad, good)
    # Anything still outside latin-1 (an accented name, say) degrades to "?"
    # rather than taking the whole export down.
    return text.encode("latin-1", "replace").decode("latin-1")


def _method_label(method: str | None) -> str:
    return (method or "—").replace("_", " ").upper()


def _date_label(value: datetime | None) -> str:
    return value.strftime("%d/%m/%Y") if value else "—"


def _rows(
    payments: list[PaymentTransaction],
    students_by_id: dict[str, Student],
    owner_names: dict[str, str],
) -> list[tuple]:
    out = []
    for p in payments:
        student = students_by_id.get(p.student_id)
        out.append(
            (
                p.receipt_number,
                student.name if student else "—",
                student.college if student else "—",
                float(p.amount),
                _method_label(p.method),
                _date_label(p.created_at),
                owner_names.get(p.owner_id, "—"),
            )
        )
    return out


def build_payments_xlsx(
    payments: list[PaymentTransaction],
    students_by_id: dict[str, Student],
    owner_names: dict[str, str],
    *,
    filter_note: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    ws.append([f"{COMPANY_NAME} — Payment ledger"])
    ws["A1"].font = Font(bold=True, size=14)
    generated = datetime.now(UTC).strftime("%d/%m/%Y %H:%M UTC")
    ws.append([f"Generated {generated}" + (f" · {filter_note}" if filter_note else "")])
    ws["A2"].font = Font(size=9, color="666666")
    ws.append([])

    # Read the row number back *after* appending rather than predicting it:
    # `append([])` does not advance max_row, so computing this up front lands
    # everything below one row too high — the totals row then overwrites the
    # last transaction and the SUM range covers the header instead.
    ws.append(_HEADERS)
    header_row = ws.max_row
    fill = PatternFill("solid", fgColor=BRAND_BLUE)
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    rows = _rows(payments, students_by_id, owner_names)
    for row in rows:
        ws.append(list(row))

    first_data_row = header_row + 1
    last_data_row = header_row + len(rows)

    # Amounts stay numeric so the column can be summed; the format is display only.
    for r in range(first_data_row, last_data_row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0.00"

    if rows:
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=3, value="Total").font = Font(bold=True)
        total = ws.cell(row=total_row, column=4)
        col = get_column_letter(4)
        total.value = f"=SUM({col}{first_data_row}:{col}{last_data_row})"
        total.font = Font(bold=True)
        total.number_format = "#,##0.00"

    widths = [16, 26, 34, 14, 16, 13, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=first_data_row, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payments_pdf(
    payments: list[PaymentTransaction],
    students_by_id: dict[str, Student],
    owner_names: dict[str, str],
    *,
    filter_note: str = "",
) -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    page_w = pdf.w

    pdf.set_fill_color(*BRAND_TEAL_RGB)
    pdf.rect(0, 0, page_w, 4, style="F")
    pdf.set_fill_color(*BRAND_BLUE_RGB)
    pdf.rect(0, 4, page_w, 16, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_xy(12, 8)
    pdf.cell(160, 8, _pdf_safe(f"{COMPANY_NAME} - Payment ledger"))

    pdf.set_text_color(*BRAND_GREY_RGB)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_xy(12, 24)
    generated = datetime.now(UTC).strftime("%d/%m/%Y %H:%M UTC")
    subtitle = f"Generated {generated}" + (f"  -  {filter_note}" if filter_note else "")
    pdf.cell(0, 5, _pdf_safe(subtitle))

    rows = _rows(payments, students_by_id, owner_names)
    widths = [30, 52, 66, 26, 26, 24, 44]

    def header() -> None:
        pdf.set_fill_color(*BRAND_BLUE_RGB)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8.5)
        for label, w in zip(_HEADERS, widths, strict=True):
            pdf.cell(w, 8, _pdf_safe(f" {label}"), border=0, fill=True)
        pdf.ln(8)

    pdf.set_y(32)
    header()

    pdf.set_text_color(*BRAND_DARK_RGB)
    pdf.set_font("Helvetica", "", 8.5)
    striped = False
    for row in rows:
        # Repeat the header when the table spills onto another page, so a
        # printed second page is still readable on its own.
        if pdf.get_y() > pdf.h - 26:
            pdf.add_page()
            pdf.set_y(16)
            header()
            pdf.set_text_color(*BRAND_DARK_RGB)
            pdf.set_font("Helvetica", "", 8.5)

        pdf.set_fill_color(245, 247, 250)
        for i, (value, w) in enumerate(zip(row, widths, strict=True)):
            text = f"{value:,.2f}" if i == 3 else str(value)
            pdf.cell(w, 7, _pdf_safe(f" {text[: int(w / 1.8)]}"), border=0, fill=striped)
        pdf.ln(7)
        striped = not striped

    if rows:
        total = sum(r[3] for r in rows)
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 9.5)
        pdf.set_text_color(*BRAND_DARK_RGB)
        pdf.cell(sum(widths[:3]), 8, _pdf_safe(f" {len(rows)} transactions"), border="T")
        pdf.cell(widths[3], 8, _pdf_safe(f" {total:,.2f}"), border="T")
        pdf.cell(sum(widths[4:]), 8, "", border="T")
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(*BRAND_GREY_RGB)
        pdf.cell(0, 10, "  No transactions match the current filters.")

    return bytes(pdf.output())
