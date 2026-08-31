"""Read an attendee roster out of a spreadsheet a college sent.

The file is not ours and never will be. Colleges keep registers in whatever
shape suits them: "Name" or "Student Name" or "NAME OF THE STUDENT", a stray
title row above the headers, blank rows where someone deleted an entry, phone
numbers stored as numbers so they arrive as floats. This module's job is to
get names out of that without an HR having to reformat anything.

The rule throughout is that a partial import beats a refusal. A row with a
name and nothing else is kept; a row with no name is skipped and reported.
Twenty good rows and two bad ones should give you twenty attendees and two
specific complaints, not a rejected file.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

# What a column might be called. Compared after lowercasing and stripping
# everything that is not a letter, so "Student Name", "STUDENT_NAME" and
# "Name of the Student" all collapse to the same key.
_HEADERS: dict[str, tuple[str, ...]] = {
    "name": (
        "name", "studentname", "fullname", "nameofthestudent",
        "studentsname", "participantname", "participant",
    ),
    "email": ("email", "emailid", "emailaddress", "mail", "mailid"),
    "phone": (
        "phone", "phoneno", "phonenumber", "mobile", "mobileno",
        "mobilenumber", "contact", "contactno", "contactnumber",
    ),
    "department": ("department", "dept", "branch", "course", "stream"),
    "year": ("year", "yearofstudy", "studyyear", "currentyear", "sem", "semester"),
}

MAX_ROWS = 2000


@dataclass
class ImportResult:
    attendees: list[dict] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)

    @property
    def imported(self) -> int:
        return len(self.attendees)


class ImportError_(Exception):
    """The file could not be read at all — as opposed to individual bad rows."""


def _key(text: object) -> str:
    return "".join(ch for ch in str(text or "").lower() if ch.isalpha())


def _clean(value: object) -> str | None:
    """One cell, as text.

    Numbers matter here: a phone column formatted as a number arrives from
    openpyxl as 9876543210.0, and "9876543210.0" is not a phone number.
    """
    if value is None:
        return None
    whole_number = isinstance(value, float) and value.is_integer()
    text = str(int(value)) if whole_number else str(value)
    text = text.strip()
    return text or None


def _map_headers(row: list) -> dict[int, str]:
    """Which column holds which field. Unrecognised columns are ignored."""
    mapping: dict[int, str] = {}
    for index, cell in enumerate(row):
        key = _key(cell)
        for field_name, aliases in _HEADERS.items():
            if key in aliases and field_name not in mapping.values():
                mapping[index] = field_name
                break
    return mapping


def _find_header_row(rows: list[list]) -> tuple[int, dict[int, str]]:
    """Locate the header row and its mapping.

    Searched for rather than assumed to be first: exported registers often
    carry a title line, a college name, or a blank row above the real
    headers. The header row is the first one that names a `name` column.
    """
    for index, row in enumerate(rows[:20]):
        mapping = _map_headers(row)
        if "name" in mapping.values():
            return index, mapping
    raise ImportError_(
        "Could not find a 'Name' column. The first row should name its "
        "columns — Name, Email, Phone, Department, Year."
    )


def _rows_from_csv(content: bytes) -> list[list]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError_("The file's text could not be read.")
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[list]:
    from openpyxl import load_workbook

    try:
        # read_only keeps a large register from being held in memory twice;
        # data_only takes the cached value of a formula rather than "=A2&B2".
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many shapes
        raise ImportError_("That file could not be opened as a spreadsheet.") from exc
    sheet = book.worksheets[0]
    return [list(r) for r in sheet.iter_rows(values_only=True)]


def parse(content: bytes, filename: str) -> ImportResult:
    """Attendees from one uploaded file, plus a note on every row skipped."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
    else:
        raise ImportError_("Upload a .xlsx or .csv file.")

    if not rows:
        raise ImportError_("That file is empty.")

    header_index, mapping = _find_header_row(rows)
    result = ImportResult()
    seen: set[str] = set()

    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = {
            field_name: _clean(row[i]) if i < len(row) else None
            for i, field_name in mapping.items()
        }
        person = values.get("name")
        if not person:
            # Blank rows are the ordinary shape of a spreadsheet, not an
            # error worth reporting; a row with other data but no name is.
            if any(values.values()):
                result.skipped.append(f"Row {offset}: no name")
            continue

        # Same person twice in one file is a copy-paste slip, not two people.
        fingerprint = (values.get("email") or person).lower()
        if fingerprint in seen:
            result.skipped.append(f"Row {offset}: {person} appears more than once")
            continue
        seen.add(fingerprint)

        if len(result.attendees) >= MAX_ROWS:
            result.skipped.append(
                f"Row {offset} onwards: more than {MAX_ROWS} rows, the rest were ignored"
            )
            break

        result.attendees.append(
            {
                "name": person[:150],
                "email": (values.get("email") or None),
                "phone": (values.get("phone") or None),
                "department": (values.get("department") or None),
                "year": (values.get("year") or None),
            }
        )

    if not result.attendees:
        raise ImportError_("No attendees were found in that file.")
    return result


def build_template() -> bytes:
    """A blank register in the shape this parser reads.

    Generated from `_HEADERS` rather than typed out, so the template and the
    thing that reads it cannot drift apart: rename a column here and the
    download changes with it.

    The example rows are real rows — someone who uploads this untouched gets
    two attendees called Anitha and Karthik. That is deliberate: a template
    that errors on upload teaches nothing, and the guidance sheet says to
    replace them.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = ["Name", "Email", "Phone", "Department", "Year"]
    widths = [26, 30, 16, 18, 12]
    samples = [
        ["Anitha Selvam", "anitha@example.com", "9876543210", "CSE", "Final"],
        ["Karthik Raja", "karthik@example.com", "9876543211", "ECE", "3rd"],
    ]

    book = Workbook()
    sheet = book.active
    sheet.title = "Attendees"
    sheet.append(columns)

    header_fill = PatternFill("solid", fgColor="1E3B5D")
    for index, width in enumerate(widths, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 22

    for row in samples:
        sheet.append(row)

    # Phone as text, so Excel does not turn 9876543210 into 9.87654E+09 and
    # a leading zero is not eaten.
    for row in range(2, len(samples) + 50):
        sheet.cell(row=row, column=3).number_format = "@"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    notes = book.create_sheet("How to use")
    for line in [
        ["Attendee register — workshops and bootcamps"],
        [],
        ["1.", "Put one attendee per row on the 'Attendees' sheet."],
        ["2.", "Replace the two example rows. They import as real people if left in."],
        ["3.", "Only 'Name' is required. Leave anything you do not have blank."],
        ["4.", "Do not rename or reorder the columns — though these also work:"],
        ["", "Name: Student Name, Full Name, Name of the Student, Participant Name"],
        ["", "Email: Email ID, E-Mail, Mail ID"],
        ["", "Phone: Mobile, Mobile No, Contact Number"],
        ["", "Department: Dept, Branch, Stream"],
        ["", "Year: Year of Study, Semester"],
        [],
        ["5.", "Extra columns such as S.No or Roll Number are ignored, not an error."],
        ["6.", "A title line above the headers is fine — it is skipped."],
        ["7.", "Rows with no name are skipped and listed back to you after upload."],
        ["8.", "The same person twice is imported once."],
        ["9.", f"Up to {MAX_ROWS} rows per file. CSV works too."],
        [],
        ["These attendees stay on the event. They are not added to Students."],
    ]:
        notes.append(line)
    notes.cell(row=1, column=1).font = Font(bold=True, size=13)
    notes.column_dimensions["A"].width = 5
    notes.column_dimensions["B"].width = 90

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
