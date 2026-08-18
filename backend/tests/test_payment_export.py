"""Payment ledger exports.

The export is a financial record, so these assert on the produced file's
contents — that every transaction survives, that the totals row sums the right
range, and that nothing is silently dropped or overwritten.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from app.models.payment import PaymentTransaction
from app.models.student import Student
from app.services.payment_export import build_payments_pdf, build_payments_xlsx


def _student(sid: str, name: str, college: str) -> Student:
    return Student(
        id=sid, application_id=None, owner_id="u1", name=name,
        email=f"{sid}@example.com", phone="9876543210", college=college,
        place="Coimbatore", category="Internship", domain="Full Stack Java",
        duration="30 Days", batch_id=None, total_fees=1, fees_paid=0,
        payment_status="paid", status="active", created_at=datetime.now(UTC),
    )


def _payment(pid: str, sid: str, receipt: str, amount: float, method: str | None):
    return PaymentTransaction(
        id=pid, student_id=sid, owner_id="u1", receipt_number=receipt,
        amount=amount, method=method, notes=None, recorded_by_id="u1",
        created_at=datetime.now(UTC),
    )


@pytest.fixture
def ledger():
    students = {
        "s1": _student("s1", "Anitha Selvam", "Kumaraguru College of Technology"),
        "s2": _student("s2", "Praveen Raghavan", "PSG College of Technology"),
        "s3": _student("s3", "Divya Krishnan", "PSG College of Technology"),
    }
    payments = [
        _payment("p1", "s1", "RCPT0001", 20000, "cash"),
        _payment("p2", "s2", "RCPT0002", 18000, "upi"),
        _payment("p3", "s3", "RCPT0003", 22000, "cash"),
    ]
    return payments, students, {"u1": "Suriya"}


def _sheet(content: bytes):
    return load_workbook(io.BytesIO(content)).active


def test_xlsx_is_a_real_workbook(ledger):
    content = build_payments_xlsx(*ledger)
    assert content[:2] == b"PK"
    assert _sheet(content).title == "Payments"


def test_every_transaction_appears_exactly_once(ledger):
    """Guards the off-by-one that let the totals row overwrite the last row."""
    payments, _, _ = ledger
    ws = _sheet(build_payments_xlsx(*ledger))
    receipts = [c.value for c in ws["A"] if isinstance(c.value, str) and c.value.startswith("RCPT")]
    assert sorted(receipts) == sorted(p.receipt_number for p in payments)


def test_totals_row_sums_only_the_data_rows(ledger):
    ws = _sheet(build_payments_xlsx(*ledger))
    header_row = next(r for r in range(1, 10) if ws.cell(row=r, column=1).value == "Receipt")
    total_row = next(r for r in range(1, 20) if ws.cell(row=r, column=3).value == "Total")

    # The totals row must sit below the data, not on top of the last entry.
    assert total_row == header_row + 3 + 1
    assert ws.cell(row=total_row, column=1).value is None

    formula = ws.cell(row=total_row, column=4).value
    assert formula == f"=SUM(D{header_row + 1}:D{header_row + 3})"


def test_amounts_are_numeric_so_the_column_can_be_summed(ledger):
    payments, _, _ = ledger
    ws = _sheet(build_payments_xlsx(*ledger))
    amounts = [
        ws.cell(row=c.row, column=4).value
        for c in ws["A"] if isinstance(c.value, str) and c.value.startswith("RCPT")
    ]
    assert all(isinstance(a, (int, float)) for a in amounts)
    assert sum(amounts) == sum(p.amount for p in payments)


def test_pdf_is_a_real_pdf(ledger):
    content = build_payments_pdf(*ledger)
    assert content[:5] == b"%PDF-"
    assert len(content) > 800


def test_exports_survive_characters_outside_latin1(ledger):
    """fpdf2's core fonts are latin-1; a typographic dash or accented name
    must not take the whole export down."""
    payments, students, owners = ledger
    students["s1"].name = "José—D’Souza"
    students["s1"].college = "Café Institute — Coimbatore"
    assert build_payments_pdf(payments, students, owners, filter_note="Café · upi")[:5] == b"%PDF-"
    assert build_payments_xlsx(payments, students, owners)[:2] == b"PK"


def test_a_deleted_student_does_not_break_the_export(ledger):
    payments, students, owners = ledger
    payments.append(_payment("p9", "gone", "RCPT0009", 5000, None))
    assert build_payments_pdf(payments, students, owners)[:5] == b"%PDF-"
    ws = _sheet(build_payments_xlsx(payments, students, owners))
    assert any(c.value == "RCPT0009" for c in ws["A"])


def test_empty_ledger_still_produces_both_files():
    assert build_payments_xlsx([], {}, {})[:2] == b"PK"
    assert build_payments_pdf([], {}, {})[:5] == b"%PDF-"
