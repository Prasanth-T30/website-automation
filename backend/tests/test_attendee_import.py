"""Reading an attendee roster out of whatever the college sent.

The parser's whole job is tolerating files nobody formatted for us, so these
tests are mostly about mess: renamed headers, a title row above the real ones,
blank rows, phone numbers stored as numbers, the same person twice.

The governing rule is that a partial import beats a refusal — twenty good rows
and two bad ones should give twenty attendees and two specific complaints.
"""

from __future__ import annotations

import io

import pytest

from app.services import attendee_import
from app.services.attendee_import import ImportError_, parse


def _csv(text: str) -> bytes:
    return text.encode("utf-8")


def _xlsx(rows: list[list]) -> bytes:
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ── the ordinary case ────────────────────────────────────────────────────


def test_a_plain_csv_imports():
    result = parse(_csv(
        "Name,Email,Phone,Department,Year\n"
        "Anitha Selvam,anitha@example.com,9876543210,CSE,Final\n"
        "Karthik Raja,karthik@example.com,9876543211,ECE,3rd\n"
    ), "roster.csv")

    assert result.imported == 2
    assert result.skipped == []
    assert result.attendees[0] == {
        "name": "Anitha Selvam", "email": "anitha@example.com",
        "phone": "9876543210", "department": "CSE", "year": "Final",
    }


def test_a_plain_xlsx_imports():
    content = _xlsx([
        ["Name", "Email", "Phone", "Department", "Year"],
        ["Anitha Selvam", "anitha@example.com", "9876543210", "CSE", "Final"],
    ])
    result = parse(content, "roster.xlsx")
    assert result.imported == 1
    assert result.attendees[0]["name"] == "Anitha Selvam"


# ── headers as colleges actually write them ──────────────────────────────


@pytest.mark.parametrize(
    "header",
    ["Name", "Student Name", "STUDENT NAME", "Full Name", "Name of the Student",
     "student_name", "Participant Name"],
)
def test_the_name_column_is_recognised_however_it_is_written(header):
    result = parse(_csv(f"{header}\nAnitha Selvam\n"), "r.csv")
    assert result.attendees[0]["name"] == "Anitha Selvam"


@pytest.mark.parametrize(
    ("header", "field_name"),
    [("Mobile No", "phone"), ("Contact Number", "phone"), ("E-Mail ID", "email"),
     ("Dept", "department"), ("Branch", "department"), ("Year of Study", "year")],
)
def test_the_other_columns_are_recognised_too(header, field_name):
    result = parse(_csv(f"Name,{header}\nAnitha,VALUE\n"), "r.csv")
    assert result.attendees[0][field_name] == "VALUE"


def test_columns_we_do_not_know_are_ignored_rather_than_fatal():
    """A register carrying a roll number or a signature column still imports."""
    result = parse(_csv(
        "S.No,Roll Number,Name,Signature\n1,21CS045,Anitha Selvam,\n"
    ), "r.csv")
    assert result.imported == 1
    assert result.attendees[0]["name"] == "Anitha Selvam"


def test_a_title_row_above_the_headers_is_skipped():
    """Exported registers routinely carry a college name or event title on
    the first line, which would otherwise be read as the header row."""
    result = parse(_csv(
        "Anna University - AI Workshop 2026\n"
        "\n"
        "Name,Email\n"
        "Anitha Selvam,anitha@example.com\n"
    ), "r.csv")
    assert result.imported == 1
    assert result.attendees[0]["name"] == "Anitha Selvam"


# ── mess inside the rows ─────────────────────────────────────────────────


def test_blank_rows_are_passed_over_silently():
    """A trailing blank line is the normal shape of a spreadsheet, not
    something to complain about."""
    result = parse(_csv("Name,Email\nAnitha,a@example.com\n\n\n"), "r.csv")
    assert result.imported == 1
    assert result.skipped == []


def test_a_row_with_data_but_no_name_is_reported():
    result = parse(_csv("Name,Email\nAnitha,a@example.com\n,orphan@example.com\n"), "r.csv")
    assert result.imported == 1
    assert len(result.skipped) == 1
    assert "no name" in result.skipped[0]


def test_the_same_person_twice_is_imported_once():
    result = parse(_csv(
        "Name,Email\n"
        "Anitha Selvam,anitha@example.com\n"
        "Anitha Selvam,anitha@example.com\n"
    ), "r.csv")
    assert result.imported == 1
    assert "more than once" in result.skipped[0]


def test_two_people_sharing_a_name_are_both_kept_when_emails_differ():
    """Common names are common. The email is what separates them."""
    result = parse(_csv(
        "Name,Email\n"
        "Anitha Selvam,anitha1@example.com\n"
        "Anitha Selvam,anitha2@example.com\n"
    ), "r.csv")
    assert result.imported == 2


def test_a_phone_stored_as_a_number_is_not_mangled():
    """Excel stores a phone column as a number, so openpyxl hands back
    9876543210.0 — and "9876543210.0" is not a phone number."""
    content = _xlsx([["Name", "Phone"], ["Anitha", 9876543210]])
    assert parse(content, "r.xlsx").attendees[0]["phone"] == "9876543210"


def test_surrounding_whitespace_is_trimmed():
    result = parse(_csv("Name,Department\n  Anitha Selvam  ,  CSE  \n"), "r.csv")
    assert result.attendees[0]["name"] == "Anitha Selvam"
    assert result.attendees[0]["department"] == "CSE"


def test_a_name_on_its_own_is_enough():
    """A register that is just a list of names is still worth importing."""
    result = parse(_csv("Name\nAnitha Selvam\nKarthik Raja\n"), "r.csv")
    assert result.imported == 2
    assert result.attendees[0]["email"] is None


# ── files that cannot be used ────────────────────────────────────────────


def test_a_file_with_no_name_column_says_so():
    with pytest.raises(ImportError_, match="Name"):
        parse(_csv("Roll,Email\n1,a@example.com\n"), "r.csv")


def test_an_empty_file_is_refused():
    with pytest.raises(ImportError_):
        parse(_csv(""), "r.csv")


def test_a_header_with_no_rows_under_it_is_refused():
    with pytest.raises(ImportError_, match="No attendees"):
        parse(_csv("Name,Email\n"), "r.csv")


def test_an_unsupported_file_type_is_refused():
    with pytest.raises(ImportError_, match="xlsx"):
        parse(_csv("Name\nAnitha\n"), "roster.pdf")


def test_something_that_is_not_a_spreadsheet_is_refused():
    with pytest.raises(ImportError_):
        parse(b"%PDF-1.4 not really a workbook", "roster.xlsx")


def test_a_huge_register_is_capped_and_says_it_was():
    rows = "Name\n" + "".join(f"Person {i}\n" for i in range(attendee_import.MAX_ROWS + 50))
    result = parse(_csv(rows), "r.csv")
    assert result.imported == attendee_import.MAX_ROWS
    assert "were ignored" in result.skipped[-1]


# ── the downloadable template ────────────────────────────────────────────


def test_the_template_imports_without_a_single_complaint():
    """The point of shipping one. If the template we hand people does not
    survive our own parser, the format is wrong somewhere."""
    result = parse(attendee_import.build_template(), "template.xlsx")
    assert result.skipped == []
    assert result.imported == 2
    assert result.attendees[0]["name"] == "Anitha Selvam"


def test_the_template_carries_every_field_the_parser_reads():
    first = parse(attendee_import.build_template(), "t.xlsx").attendees[0]
    assert all(first[f] for f in ("name", "email", "phone", "department", "year"))


def test_the_template_phone_survives_excels_number_formatting():
    """The column is written as text precisely so 9876543210 does not come
    back as 9.87654E+09 or lose a leading zero."""
    assert parse(attendee_import.build_template(), "t.xlsx").attendees[0]["phone"] == (
        "9876543210"
    )


def test_the_guidance_lives_on_a_second_sheet_the_parser_never_reads():
    """Instructions on the first sheet would be imported as attendees."""
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(attendee_import.build_template()))
    assert book.sheetnames == ["Attendees", "How to use"]
    assert book.worksheets[0]["A1"].value == "Name"
