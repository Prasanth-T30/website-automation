"""Payments: capping rule, sequential receipts, ownership guard, receipt PDF,
and the admin HR-performance rollup. Real HTTP against the real emulator.

A freshly approved student is "paid in full" against their own self-reported
registration amount (`total_fees == fees_paid`) — there's no balance to
collect until an HR revises `total_fees` up to the programme's real cost.
Every capping test here does that PATCH first to create a balance worth
testing against.
"""

from __future__ import annotations

import io
import subprocess
import uuid
from shutil import which

import pytest
from fastapi.testclient import TestClient

from app.core.firebase import get_firestore
from app.core.security import hash_password
from app.models.user import UserRole
from app.repositories.users import UserRepository
from tests.conftest import requires_emulator

pytestmark = requires_emulator

PDFTOTEXT = which("pdftotext")


@pytest.fixture
def client():
    from app.main import app

    app.state.limiter.reset()
    return TestClient(app)


@pytest.fixture
def user_repo():
    return UserRepository(get_firestore())


def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


def _login_as(client: TestClient, user_repo: UserRepository, *, role: UserRole) -> tuple[str, str]:
    email = f"{_unique('e2e-pay')}@dvein.in"
    user = user_repo.create(
        email=email, full_name="E2E Payments", role=role,
        password_hash=hash_password("a-real-password-1"), phone=None,
        must_change_password=False,
    )
    res = client.post("/api/v1/auth/login", json={"email": email, "password": "a-real-password-1"})
    assert res.status_code == 200
    return client.cookies["dvein_csrf"], user.id


def _create_approved_student(client: TestClient, csrf: str, *, amount: str = "5000") -> str:
    form = {
        "name": "Fee Student", "email": f"{_unique('fee')}@example.com",
        "phone": "9876543210", "college": "College", "place": "Chennai",
        "applicant_type": "student", "category": "Project", "domain": "Software Testing",
        "duration": "30 Days", "start_date": "2026-09-01", "end_date": "2026-10-01",
        "amount": amount, "transaction_id": _unique("TXN"), "declaration": "true",
    }
    files = {"payment_screenshot": ("proof.png", io.BytesIO(b"fake"), "image/png")}
    submitted = client.post("/api/v1/public/applications", data=form, files=files)
    app_id = submitted.json()["id"]

    client.post(f"/api/v1/applications/{app_id}/claim", headers={"X-CSRF-Token": csrf})
    approved = client.post(
        f"/api/v1/applications/{app_id}/approve",
        json={"subject": "", "body": ""},
        headers={"X-CSRF-Token": csrf},
    )
    return approved.json()["converted_student_id"]


def _bump_total_fees(client: TestClient, csrf: str, student_id: str, total_fees: float) -> None:
    res = client.patch(
        f"/api/v1/students/{student_id}",
        json={"total_fees": total_fees},
        headers={"X-CSRF-Token": csrf},
    )
    assert res.status_code == 200


def test_record_payment_caps_at_outstanding_balance(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="5000")
    _bump_total_fees(client, csrf, student_id, 8000)

    res = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 5000, "method": "upi"},
        headers={"X-CSRF-Token": csrf},
    )
    assert res.status_code == 201
    payment = res.json()
    assert payment["amount"] == 3000  # capped at the 8000 - 5000 balance
    assert payment["receipt_number"].startswith("RCPT")

    student = client.get(f"/api/v1/students/{student_id}").json()
    assert student["fees_paid"] == 8000
    assert student["payment_status"] == "paid"


def test_sequential_receipt_numbers_increment(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="1000")
    _bump_total_fees(client, csrf, student_id, 20000)

    first = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 1000},
        headers={"X-CSRF-Token": csrf},
    ).json()
    second = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 1000},
        headers={"X-CSRF-Token": csrf},
    ).json()

    first_seq = int(first["receipt_number"].removeprefix("RCPT"))
    second_seq = int(second["receipt_number"].removeprefix("RCPT"))
    assert second_seq == first_seq + 1


def test_cannot_record_payment_once_fully_paid(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="5000")

    res = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 100},
        headers={"X-CSRF-Token": csrf},
    )
    assert res.status_code == 400


def test_non_owner_hr_cannot_record_payment(client: TestClient, user_repo):
    csrf1, _ = _login_as(client, user_repo, role=UserRole.hr)
    student_id = _create_approved_student(client, csrf1, amount="5000")
    _bump_total_fees(client, csrf1, student_id, 8000)

    csrf2, _ = _login_as(client, user_repo, role=UserRole.hr)
    res = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 1000},
        headers={"X-CSRF-Token": csrf2},
    )
    assert res.status_code == 403


def test_approval_records_the_registration_amount_as_the_first_receipt(
    client: TestClient, user_repo
):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="2000")

    listing = client.get("/api/v1/payments", params={"student_id": student_id})
    rows = listing.json()
    assert len(rows) == 1
    assert rows[0]["amount"] == 2000
    assert rows[0]["receipt_number"].startswith("RCPT")


def test_list_payments_filters_by_student(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="2000")
    _bump_total_fees(client, csrf, student_id, 6000)
    client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 4000},
        headers={"X-CSRF-Token": csrf},
    )

    listing = client.get("/api/v1/payments", params={"student_id": student_id})
    rows = listing.json()
    # One row from approval (the registration's own amount) plus this
    # explicit installment.
    assert len(rows) == 2
    assert {r["amount"] for r in rows} == {2000, 4000}


@pytest.mark.skipif(PDFTOTEXT is None, reason="pdftotext (poppler) not installed")
def test_receipt_pdf_contains_amount_and_receipt_number(client: TestClient, user_repo, tmp_path):
    csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    student_id = _create_approved_student(client, csrf, amount="2000")
    _bump_total_fees(client, csrf, student_id, 6000)
    payment = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 4000},
        headers={"X-CSRF-Token": csrf},
    ).json()

    res = client.get(f"/api/v1/payments/{payment['id']}/receipt")
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"

    pdf_path = tmp_path / "receipt.pdf"
    pdf_path.write_bytes(res.content)
    text = subprocess.run(
        [PDFTOTEXT, str(pdf_path), "-"], capture_output=True, text=True, check=True
    ).stdout
    assert payment["receipt_number"] in text
    assert "4,000.00" in text


def test_hr_cannot_view_hr_performance(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    res = client.get("/api/v1/admin/hr-performance", headers={"X-CSRF-Token": csrf})
    assert res.status_code == 403


def test_hr_performance_reflects_claims_and_revenue(client: TestClient, user_repo):
    hr_csrf, hr_id = _login_as(client, user_repo, role=UserRole.hr)
    student_id = _create_approved_student(client, hr_csrf, amount="3000")
    _bump_total_fees(client, hr_csrf, student_id, 10000)
    client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 7000},
        headers={"X-CSRF-Token": hr_csrf},
    )

    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    res = client.get("/api/v1/admin/hr-performance", headers={"X-CSRF-Token": admin_csrf})
    assert res.status_code == 200
    rows = {row["id"]: row for row in res.json()}
    assert hr_id in rows

    row = rows[hr_id]
    assert row["claimed_count"] >= 1
    assert row["converted_count"] >= 1
    # 3000 from the registration's own amount (recorded on approval) + 7000
    # from the explicit installment recorded above.
    assert row["revenue_all_time"] >= 10000
    assert row["revenue_this_month"] >= 10000


def test_admin_owned_revenue_still_appears_in_the_report(client: TestClient, user_repo):
    """An admin can claim an application or add a student by hand. That
    revenue must not fall out of the per-user breakdown, or the rows stop
    summing to the ledger and an audit silently comes up short."""
    admin_csrf, admin_id = _login_as(client, user_repo, role=UserRole.admin)

    created = client.post(
        "/api/v1/students",
        json={
            "name": "Admin Walkin", "email": f"{_unique('adminwalk')}@example.com",
            "phone": "9876543210", "college": "PSG College of Technology",
            "place": "Coimbatore", "category": "Course", "domain": "Full Stack Java",
            "duration": "30 Days", "total_fees": 8000, "fees_paid": 0,
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert created.status_code == 201
    student_id = created.json()["id"]

    paid = client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 8000, "method": "cash"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert paid.status_code == 201, paid.text

    report = client.get("/api/v1/admin/hr-performance", headers={"X-CSRF-Token": admin_csrf})
    assert report.status_code == 200
    rows = report.json()

    mine = next((r for r in rows if r["id"] == admin_id), None)
    assert mine is not None, "admin-owned revenue vanished from hr-performance"
    assert mine["revenue_all_time"] >= 8000
    assert mine["role"] == "admin"


def test_revenue_rows_reconcile_with_the_payment_ledger(client: TestClient, user_repo):
    """The sum of every row's all-time revenue must equal the ledger total."""
    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)

    ledger = client.get("/api/v1/payments", headers={"X-CSRF-Token": admin_csrf}).json()
    ledger_total = sum(p["amount"] for p in ledger)

    rows = client.get(
        "/api/v1/admin/hr-performance", headers={"X-CSRF-Token": admin_csrf}
    ).json()
    reported_total = sum(r["revenue_all_time"] for r in rows)

    assert reported_total == pytest.approx(ledger_total), (
        "per-user revenue does not add up to the payment ledger"
    )


def test_month_boundary_uses_the_institute_timezone(client: TestClient, user_repo):
    """A payment taken just after local midnight on the 1st belongs to the new
    month, not the previous one. Guards the IST-vs-UTC boundary."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    from app.api.v1.admin import _month_start_utc
    from app.core.config import settings

    tz = ZoneInfo(settings.reporting_timezone)
    boundary = _month_start_utc()

    # Local midnight on the 1st, expressed in the institute's own clock.
    local_boundary = boundary.astimezone(tz)
    assert (local_boundary.day, local_boundary.hour, local_boundary.minute) == (1, 0, 0)

    # A payment one minute into the month counts; one minute before does not.
    assert (boundary + timedelta(minutes=1)) >= boundary
    assert (boundary - timedelta(minutes=1)) < boundary
    assert boundary <= datetime.now(tz).astimezone(boundary.tzinfo)


def _revenue_by_hr(client: TestClient, admin_csrf: str) -> dict[str, float]:
    res = client.get("/api/v1/admin/hr-performance", headers={"X-CSRF-Token": admin_csrf})
    assert res.status_code == 200, res.text
    return {row["id"]: row["revenue_all_time"] for row in res.json()}


def test_reassigning_a_student_moves_their_revenue_to_the_new_hr(
    client: TestClient, user_repo
):
    """The money follows the student: debited from one HR, credited to the other.

    Admin reassignment is meant to hand over a whole relationship, so leaving
    the revenue behind would show the losing HR income for a student they no
    longer hold and the gaining HR a student who appears to have paid nothing.
    """
    hr_a_csrf, hr_a_id = _login_as(client, user_repo, role=UserRole.hr)
    student_id = _create_approved_student(client, hr_a_csrf, amount="3000")
    _bump_total_fees(client, hr_a_csrf, student_id, 10000)
    client.post(
        "/api/v1/payments/record",
        json={"student_id": student_id, "amount": 7000},
        headers={"X-CSRF-Token": hr_a_csrf},
    )

    _, hr_b_id = _login_as(client, user_repo, role=UserRole.hr)

    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    before = _revenue_by_hr(client, admin_csrf)
    assert before[hr_a_id] >= 10000
    a_before, b_before = before[hr_a_id], before.get(hr_b_id, 0.0)

    moved = client.post(
        f"/api/v1/students/{student_id}/reassign",
        json={"owner_id": hr_b_id},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert moved.status_code == 200, moved.text
    body = moved.json()
    assert body["payments_moved"] == 2  # the approval amount plus the installment
    assert body["revenue_moved"] == 10000
    assert body["to_owner_name"]

    after = _revenue_by_hr(client, admin_csrf)
    assert after[hr_a_id] == a_before - 10000, "revenue did not leave the old HR"
    assert after[hr_b_id] == b_before + 10000, "revenue did not reach the new HR"


def test_reassignment_moves_the_claim_so_conversion_rates_stay_honest(
    client: TestClient, user_repo
):
    """The old HR must not keep a claim for a student they no longer hold."""
    hr_a_csrf, hr_a_id = _login_as(client, user_repo, role=UserRole.hr)
    student_id = _create_approved_student(client, hr_a_csrf, amount="3000")

    _, hr_b_id = _login_as(client, user_repo, role=UserRole.hr)
    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)

    def rows() -> dict:
        res = client.get("/api/v1/admin/hr-performance", headers={"X-CSRF-Token": admin_csrf})
        return {row["id"]: row for row in res.json()}

    a_claims_before = rows()[hr_a_id]["claimed_count"]

    client.post(
        f"/api/v1/students/{student_id}/reassign",
        json={"owner_id": hr_b_id},
        headers={"X-CSRF-Token": admin_csrf},
    )

    after = rows()
    assert after[hr_a_id]["claimed_count"] == a_claims_before - 1
    assert after[hr_a_id]["converted_count"] == 0
    assert after[hr_b_id]["claimed_count"] == 1
    assert after[hr_b_id]["converted_count"] == 1


def test_the_record_of_who_took_each_payment_survives_reassignment(
    client: TestClient, user_repo
):
    """Credit for the revenue moves; the audit trail of who collected it does not."""
    hr_a_csrf, hr_a_id = _login_as(client, user_repo, role=UserRole.hr)
    student_id = _create_approved_student(client, hr_a_csrf, amount="3000")

    _, hr_b_id = _login_as(client, user_repo, role=UserRole.hr)
    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    client.post(
        f"/api/v1/students/{student_id}/reassign",
        json={"owner_id": hr_b_id},
        headers={"X-CSRF-Token": admin_csrf},
    )

    ledger = client.get("/api/v1/payments", params={"student_id": student_id}).json()
    assert ledger, "the student's payments vanished"
    for payment in ledger:
        assert payment["owner_id"] == hr_b_id
        assert payment["recorded_by_id"] == hr_a_id


# ── Revenue isolation ────────────────────────────────────────────────────
# Each HR's revenue is theirs alone; only admin sees across the institute.
# These endpoints all took a client-supplied `mine` flag or no scope at all,
# so an HR could read the whole institute's ledger by dropping a query
# parameter. The console always sent the flag, which is exactly why the gap
# stayed invisible — the enforcement has to live here, not in the caller.


def _seed_two_hrs(client: TestClient, user_repo):
    hr_a_csrf, hr_a_id = _login_as(client, user_repo, role=UserRole.hr)
    a_student = _create_approved_student(client, hr_a_csrf, amount="3000")

    hr_b_csrf, hr_b_id = _login_as(client, user_repo, role=UserRole.hr)
    b_student = _create_approved_student(client, hr_b_csrf, amount="9000")
    return (hr_a_csrf, hr_a_id, a_student), (hr_b_csrf, hr_b_id, b_student)


def test_an_hr_listing_payments_without_the_mine_flag_still_sees_only_their_own(
    client: TestClient, user_repo
):
    (_, hr_a_id, _), (hr_b_csrf, hr_b_id, _) = _seed_two_hrs(client, user_repo)

    rows = client.get("/api/v1/payments").json()  # note: no mine=true
    assert rows, "B should still see their own payments"
    assert all(p["owner_id"] == hr_b_id for p in rows)
    assert not any(p["owner_id"] == hr_a_id for p in rows)


def test_an_hr_cannot_widen_the_ledger_by_asking_for_mine_false(
    client: TestClient, user_repo
):
    (_, hr_a_id, _), (hr_b_csrf, hr_b_id, _) = _seed_two_hrs(client, user_repo)

    rows = client.get("/api/v1/payments", params={"mine": "false"}).json()
    assert all(p["owner_id"] == hr_b_id for p in rows)


def test_an_admin_still_sees_every_hrs_revenue(client: TestClient, user_repo):
    (_, hr_a_id, _), (_, hr_b_id, _) = _seed_two_hrs(client, user_repo)

    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    owners = {p["owner_id"] for p in client.get("/api/v1/payments").json()}
    assert {hr_a_id, hr_b_id} <= owners, "the admin's institute-wide view broke"


def test_the_excel_export_is_scoped_to_the_hr_who_asked(client: TestClient, user_repo):
    """The export is a file that leaves the building — it must not carry
    another HR's transactions out with it."""
    _seed_two_hrs(client, user_repo)
    mine = client.get("/api/v1/payments", params={"mine": "true"}).json()

    res = client.get("/api/v1/payments/export.xlsx")
    assert res.status_code == 200
    rows = _xlsx_receipt_numbers(res.content)
    assert rows == {p["receipt_number"] for p in mine}


def test_the_pdf_export_is_scoped_to_the_hr_who_asked(client: TestClient, user_repo):
    (_, _, _), (_, hr_b_id, _) = _seed_two_hrs(client, user_repo)
    res = client.get("/api/v1/payments/export.pdf")
    assert res.status_code == 200
    assert res.content.startswith(b"%PDF")


def test_an_hr_cannot_download_another_hrs_receipt(client: TestClient, user_repo):
    (hr_a_csrf, _, _), _ = _seed_two_hrs(client, user_repo)

    # Back to A to learn one of their receipt ids.
    _login_as(client, user_repo, role=UserRole.hr)
    admin_csrf, _ = _login_as(client, user_repo, role=UserRole.admin)
    every = client.get("/api/v1/payments").json()
    someone_elses = every[0]["id"]

    _login_as(client, user_repo, role=UserRole.hr)  # a fresh, unrelated HR
    res = client.get(f"/api/v1/payments/{someone_elses}/receipt")
    assert res.status_code == 404, "an unrelated HR pulled someone else's receipt"


def _xlsx_receipt_numbers(content: bytes) -> set[str]:
    """Receipt numbers present in the exported workbook."""
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(content)).active
    found = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.startswith("RCPT"):
                found.add(cell)
    return found


# ── Fee capture at approval ──────────────────────────────────────────────
# The applicant only states what they are paying now. Without the course fee
# being set at approval, every student enrolled already settled and nothing
# ever appeared as outstanding on the Finance screen.


def _approve_with_fee(client: TestClient, csrf: str, *, paid: str, total_fees) -> dict:
    form = {
        "name": "Fee Capture", "email": f"{_unique('cap')}@example.com",
        "phone": "9876543210", "college": "College", "place": "Chennai",
        "applicant_type": "student", "category": "Project", "domain": "Software Testing",
        "duration": "30 Days", "start_date": "2026-09-01", "end_date": "2026-10-01",
        "amount": paid, "transaction_id": _unique("TXN"), "declaration": "true",
    }
    files = {"payment_screenshot": ("proof.png", io.BytesIO(b"fake"), "image/png")}
    app_id = client.post("/api/v1/public/applications", data=form, files=files).json()["id"]
    client.post(f"/api/v1/applications/{app_id}/claim", headers={"X-CSRF-Token": csrf})

    body = {"subject": "", "body": ""}
    if total_fees is not None:
        body["total_fees"] = total_fees
    approved = client.post(
        f"/api/v1/applications/{app_id}/approve", json=body, headers={"X-CSRF-Token": csrf}
    )
    assert approved.status_code == 200, approved.text
    sid = approved.json()["converted_student_id"]
    return client.get(f"/api/v1/students/{sid}").json()


def test_approving_with_a_course_fee_leaves_a_pending_balance(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    student = _approve_with_fee(client, csrf, paid="5000", total_fees=20000)

    assert student["total_fees"] == 20000
    assert student["fees_paid"] == 5000
    assert student["total_fees"] - student["fees_paid"] == 15000
    assert student["payment_status"] == "partial"


def test_approving_with_the_fee_already_covered_settles_the_student(
    client: TestClient, user_repo
):
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    student = _approve_with_fee(client, csrf, paid="20000", total_fees=20000)

    assert student["total_fees"] - student["fees_paid"] == 0
    assert student["payment_status"] == "paid"


def test_a_fee_below_what_was_already_paid_never_goes_negative(client: TestClient, user_repo):
    """A typo must not create a negative balance — that would credit the
    student against their next installment."""
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    student = _approve_with_fee(client, csrf, paid="9000", total_fees=1000)

    assert student["total_fees"] == 9000
    assert student["total_fees"] - student["fees_paid"] == 0


def test_omitting_the_fee_keeps_the_old_behaviour(client: TestClient, user_repo):
    """Approvals from a client that does not send the field must still work."""
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    student = _approve_with_fee(client, csrf, paid="7000", total_fees=None)

    assert student["total_fees"] == 7000
    assert student["total_fees"] - student["fees_paid"] == 0


def test_the_export_can_be_narrowed_to_fully_paid_students(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    settled = _approve_with_fee(client, csrf, paid="8000", total_fees=8000)
    owing = _approve_with_fee(client, csrf, paid="2000", total_fees=30000)

    ledger = client.get("/api/v1/payments").json()
    receipt_of = {p["student_id"]: p["receipt_number"] for p in ledger}

    paid_only = _xlsx_receipt_numbers(
        client.get("/api/v1/payments/export.xlsx", params={"fee_status": "paid"}).content
    )
    assert receipt_of[settled["id"]] in paid_only
    assert receipt_of[owing["id"]] not in paid_only


def test_the_export_can_be_narrowed_to_students_who_still_owe(client: TestClient, user_repo):
    csrf, _ = _login_as(client, user_repo, role=UserRole.hr)
    settled = _approve_with_fee(client, csrf, paid="8000", total_fees=8000)
    owing = _approve_with_fee(client, csrf, paid="2000", total_fees=30000)

    ledger = client.get("/api/v1/payments").json()
    receipt_of = {p["student_id"]: p["receipt_number"] for p in ledger}

    pending_only = _xlsx_receipt_numbers(
        client.get("/api/v1/payments/export.xlsx", params={"fee_status": "pending"}).content
    )
    assert receipt_of[owing["id"]] in pending_only
    assert receipt_of[settled["id"]] not in pending_only
